# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl"]
# ///
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
import warnings

try:
    from .resources import sentence_pattern_list, irodori_can_do
except ImportError:  # running as a standalone script
    from resources import sentence_pattern_list, irodori_can_do

warnings.filterwarnings('ignore', message='Print area cannot be set', category=UserWarning)

groupings = {
    'sheet_sentences': {
        '冊': 'book',
        'トピック': 'topic',
        'タイトル': 'lesson',
        '文型': 'sentence_pattern',
        '例文': 'sentence_example'
    },
    'sheet_can_do': {
        '冊': 'book',
        'トピック': 'topic',
        '課': 'lesson',
        '活動タイトル': 'skill_title',
        '技能': 'skill_type',
        'Can-do': 'skill_description',
        '元にした生活日本語Can-do_2': 'skill_tags'
    }
}


def _is_colored(font):
    color = getattr(font, 'color', None)
    rgb = getattr(color, 'rgb', None)
    return isinstance(rgb, str) and rgb[-6:].upper() != '000000'


def get_highlighted(cell: CellRichText | None, rep=' '):
    if cell is None:
        return None, []

    def is_colored(font):
        color = getattr(font, 'color', None)
        rgb = getattr(color, 'rgb', None)
        return isinstance(rgb, str) and rgb[-6:].upper() != '000000'

    spans = [(p.text, is_colored(p.font)) if isinstance(p, TextBlock) else (p, False) for p in cell]
    raw = ''.join(t for t, _ in spans).replace('\n', rep)
    lead = len(raw) - len(raw.lstrip())
    clean = raw.strip()

    ranges = []
    offset = 0
    for text, fmt in spans:
        n = len(text.replace('\n', rep))
        if fmt and n:
            s = max(offset - lead, 0)
            e = min(offset + n - lead, len(clean))
            if s < e:
                if ranges and ranges[-1][1] == s:
                    ranges[-1][1] = e
                else:
                    ranges.append([s, e])
        offset += n
    return clean, ranges


def read_xlsx(path: Path, name: str, header_row: int, limit: int = 0) -> list[dict]:
    wb = openpyxl.load_workbook(path, rich_text=True)
    ws = wb.active

    # map every coordinate inside a merged range to its anchor (top-left) value
    cells = {}
    for rng in ws.merged_cells.ranges:
        anchor = ws.cell(row=rng.min_row, column=rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                cells[(r, c)] = anchor

    def get_cell(cell, rep=' ', as_rich_text: bool = False) -> str | CellRichText | None:
        v = cells.get((cell.row, cell.column), cell.value)
        if v is None:
            return None
        if isinstance(v, CellRichText):
            if as_rich_text:
                return v
            v = ''.join(p.text if isinstance(p, TextBlock) else p for p in v)
        else:
            v = str(v)
        return v.replace('\n', rep).strip()

    def get_headers():
        raw_headers = [get_cell(c, rep='') for c in next(rows)]
        seen = {}
        headers = []
        for h in raw_headers:
            if h is None:
                headers.append(None)
                continue
            seen[h] = seen.get(h, 0) + 1
            headers.append(h if seen[h] == 1 else f'{h}_{seen[h]}')
        return headers


    rows = ws.iter_rows(min_row=header_row)
    headers = get_headers()

    results = []
    for row in rows:
        entry = {}
        for header, cell in zip(headers, row):
            if header is None:
                continue
            if name == 'sheet_sentences' and header == '例文':
                cell = get_cell(cell, as_rich_text=True)
                cell, highlights = get_highlighted(cell)
                entry[header] = cell
                entry['*sentence_highlights'] = highlights
            else:
                entry[header] = get_cell(cell)
        results.append(entry)

        if limit > 0 and len(results) >= limit:
            break

    headers = groupings[name]
    results_updated = [{} for _ in range(len(results))]
    for i in range(len(results)):
        result = results_updated[i]
        result.update({headers[k]: v for k, v in results[i].items() if k in headers})
        result.update({k[1:]: v for k, v in results[i].items() if k.startswith('*')})
        # results_updated[i] = result

    wb.close()
    return results_updated


def input_data() -> dict[int, dict]:
    files = [
        (irodori_can_do, 'sheet_can_do', 3),
        (sentence_pattern_list, 'sheet_sentences', 3),
    ]

    results = [read_xlsx(resource.path, name, header_row) for resource, name, header_row in files]
    headers = [set(result[0].keys()) for result in results]
    headers_i = set.intersection(*headers)

    items = {}
    for result in results:
        for entry in result:
            try:
                key = '_'.join([entry[header] for header in headers_i])
                if key not in items:
                    items[key] = {}

                if 'skill_tags' in entry:
                    tags = entry.pop('skill_tags').split('／')
                    entry['skill_cefr'] = tags[0]
                    entry['skill_activity'] = tags[1]
                    entry['skill_situation'] = tags[2]

                items[key].update({k: v for k, v in entry.items()})
            except Exception:
                continue

    data = {i: item for i, item in enumerate(items.values())}

    claude_input = Path(__file__).resolve().parent / 'resources' / 'claude_input.json'
    claude_input.write_text(json.dumps([{'index': k, **v} for k, v in data.items()], ensure_ascii=False, indent=2, default=str), encoding='utf-8')

    return data
